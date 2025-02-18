import os
from werkzeug.utils import secure_filename

from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from datetime import datetime
import base64
import json
from models import db, Task, JournalEntry, Contact
from config import Config

app = Flask(__name__)

@app.template_filter('from_json')
def from_json(value):
    import json
    return json.loads(value) if value else []
app.config.from_object(Config)

db.init_app(app)

from api import api
app.register_blueprint(api)

# Создаем таблицы, если они не существуют
with app.app_context():
    db.create_all()

@app.route('/')
def home():
    tasks = Task.query.order_by(Task.due_date.asc()).all()
    today = datetime.now().strftime('%Y-%m-%d')

    # Подсчет статистики
    total_tasks = Task.query.count()  # Общее количество задач
    completed_tasks = Task.query.filter_by(completed=True).count()
    important_tasks = Task.query.filter_by(important=True).count()
    urgent_tasks = Task.query.filter_by(urgent=True).count()
    regular_tasks = Task.query.filter_by(important=False, urgent=False).count()  # Обычные задачи



    # Получаем последние действия
    recent_activities = [
        {'icon': 'fas fa-check', 'text': 'Задача "' + t.title + '" выполнена', 
         'time': t.created_date.strftime('%H:%M')}
        for t in Task.query.filter_by(completed=True).order_by(Task.created_date.desc()).limit(5)
    ]

    # Получаем предстоящие задачи
    upcoming_tasks = Task.query.filter(
        Task.completed == False,
        Task.due_date >= datetime.now()
    ).order_by(Task.due_date.asc()).limit(5)

    return render_template('index.html', 
        tasks=tasks,
        today=today,
        stats={
            'total': total_tasks,
            'regular': regular_tasks,
            'completed': completed_tasks,
            'important': important_tasks,
            'urgent': urgent_tasks
        },
        upcoming_tasks=upcoming_tasks,
        recent_activities=recent_activities
    )

@app.route('/calendar')
def calendar():
    tasks = Task.query.all()
    calendar_tasks = []
    for task in tasks:
        calendar_tasks.append({
            'id': task.id,
            'title': task.title,
            'start': task.due_date.strftime('%Y-%m-%d') if task.due_date else None,
            'description': task.description,
            'completed': task.completed,
            'important': task.important,
            'urgent': task.urgent,
            'className': 'completed' if task.completed else ''
        })
    return render_template('calendar.html', tasks=calendar_tasks)

@app.route('/get_task/<int:id>')
def get_task(id):
    task = Task.query.get_or_404(id)
    return {
        'id': task.id,
        'title': task.title,
        'description': task.description,
        'due_date': task.due_date.strftime('%Y-%m-%d') if task.due_date else None,
        'completed': task.completed,
        'important': task.important,
        'urgent': task.urgent
    }

@app.route('/tasks')
def tasks():
    tasks = Task.query.order_by(Task.due_date.asc()).all()
    today = datetime.now().strftime('%Y-%m-%d')
    return render_template('tasks.html', tasks=tasks, today=today)

@app.route('/add_task', methods=['POST'])
def add_task():
    try:
        title = request.form['title']
        description = request.form['description']
        due_date = request.form.get('due_date')
        important = request.form.get('important') == 'true'
        urgent = request.form.get('urgent') == 'true'

        if not title:
            flash('Название задачи обязательно')
            return redirect(url_for('calendar'))

        new_task = Task(
            title=title,
            description=description,
            due_date=datetime.strptime(due_date, '%Y-%m-%d') if due_date else None,
            important=important,
            urgent=urgent
        )
        db.session.add(new_task)
        db.session.commit()
        return redirect(url_for('calendar'))
    except Exception as e:
        db.session.rollback()
        flash('Произошла ошибка при добавлении задачи')
        return redirect(url_for('calendar'))

@app.route('/complete_task/<int:id>', methods=['POST'])
def complete_task(id):
    task = Task.query.get_or_404(id)
    task.completed = not task.completed
    db.session.commit()
    return {'success': True, 'completed': task.completed}

@app.route('/toggle_important/<int:id>', methods=['POST'])
def toggle_important(id):
    task = Task.query.get_or_404(id)
    task.important = not task.important
    db.session.commit()
    return {'success': True, 'important': task.important}

@app.route('/toggle_urgent/<int:id>', methods=['POST'])
def toggle_urgent(id):
    task = Task.query.get_or_404(id)
    task.urgent = not task.urgent
    db.session.commit()
    return {'success': True, 'urgent': task.urgent}


@app.route('/export_contacts')
def export_contacts():
    from openpyxl import Workbook
    import json
    from io import BytesIO
    from flask import send_file
    
    contacts = Contact.query.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Контакты"
    
    # Заголовки
    headers = ['Имя', 'Должность', 'Email', 'Телефоны', 'Организация', 'Заметки']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Данные
    for row, contact in enumerate(contacts, 2):
        phones = json.loads(contact.phones) if contact.phones else []
        phones_str = ', '.join(phones)
        
        ws.cell(row=row, column=1, value=contact.name)
        ws.cell(row=row, column=2, value=contact.position)
        ws.cell(row=row, column=3, value=contact.email)
        ws.cell(row=row, column=4, value=phones_str)
        ws.cell(row=row, column=5, value=contact.department)
        ws.cell(row=row, column=6, value=contact.notes)
    
    # Автонастройка ширины столбцов
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Сохранение в буфер
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='contacts.xlsx'
    )

@app.route('/delete_task/<int:id>', methods=['POST'])
def delete_task(id):
    task = Task.query.get_or_404(id)
    db.session.delete(task)
    db.session.commit()
    return {'success': True}

@app.route('/edit_task', methods=['POST'])
def edit_task():
    task_id = request.form['task_id']
    task = Task.query.get_or_404(task_id)
    task.title = request.form['title']
    task.description = request.form['description']
    task.due_date = datetime.strptime(request.form['due_date'], '%Y-%m-%d')
    task.important = request.form.get('important') == 'true'
    task.urgent = request.form.get('urgent') == 'true'
    db.session.commit()
    return jsonify({
        'success': True,
        'task': {
            'id': task.id,
            'title': task.title,
            'description': task.description,
            'due_date': task.due_date.strftime('%Y-%m-%d'),
            'completed': task.completed,
            'important': task.important,
            'urgent': task.urgent
        }
    })

@app.route('/journal')
def journal():
    entries = JournalEntry.query.order_by(JournalEntry.created_date.desc()).all()
    return render_template('journal.html', entries=entries)

@app.route('/add_journal_entry', methods=['POST'])
def add_journal_entry():
    entry = JournalEntry(
        title=request.form['title'],
        content=request.form['content'],
        entry_type=request.form['entry_type']
    )
    db.session.add(entry)
    db.session.commit()
    return redirect(url_for('journal'))

@app.route('/contacts')
def contacts():
    contacts = Contact.query.all()
    departments = {contact.department for contact in contacts if contact.department}
    return render_template('contacts.html', contacts=contacts, departments=departments)

def process_image(image_data, max_size=(800, 800), quality=85):
    from io import BytesIO
    from PIL import Image
    import imghdr
    
    if not image_data:
        return None
        
    # Validate image type
    image_type = imghdr.what(None, image_data)
    if image_type not in ['jpeg', 'png', 'gif']:
        raise ValueError('Invalid image type')
        
    try:
        image = Image.open(BytesIO(image_data))
        
        # Convert to RGB if necessary
        if image.mode not in ('L', 'RGB'):
            image = image.convert('RGB')
            
        # Resize if larger than max_size
        if image.size[0] > max_size[0] or image.size[1] > max_size[1]:
            image.thumbnail(max_size, Image.LANCZOS)
            
        # Save optimized image
        output = BytesIO()
        image.save(output, format='JPEG', quality=quality, optimize=True)
        return output.getvalue()
    except Exception as e:
        raise ValueError(f'Image processing error: {str(e)}')

@app.route('/add_contact', methods=['POST'])
def add_contact():
    try:
        photo = request.files.get('photo')
        photo_filename = None

        # Создаем директорию, если её нет
        upload_dir = os.path.join(app.root_path, 'static/uploads/contacts')
        os.makedirs(upload_dir, exist_ok=True)

        if photo and photo.filename:
            try:
                # Generate unique filename
                filename = secure_filename(photo.filename)
                photo_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
                photo_path = os.path.join(upload_dir, photo_filename)

                # Process and save image
                photo_data = photo.read()
                if photo_data:
                    from PIL import Image
                    from io import BytesIO
                    img = Image.open(BytesIO(photo_data))
                    img.save(photo_path, optimize=True, quality=85)
            except Exception as e:
                print(f"Error processing image: {str(e)}")
                photo_filename = None

        import json
        phones = request.form.getlist('phones[]')
        # Фильтруем пустые значения и сохраняем как JSON
        phones = [phone for phone in phones if phone.strip()]
        contact = Contact(
            name=request.form['name'],
            position=request.form['position'],
            email=request.form['email'],
            phones=json.dumps(phones),  # Конвертируем список в JSON строку
            department=request.form['department'],
            notes=request.form.get('notes'),
            photo_path=photo_filename
        )
        db.session.add(contact)
        db.session.commit()

        return jsonify({
            'success': True,
            'contact': {
                'id': contact.id,
                'name': contact.name,
                'position': contact.position,
                'email': contact.email,
                'phones': json.loads(contact.phones),
                'department': contact.department,
                'photo_url': url_for('static', filename=f'uploads/contacts/{photo_filename}') if photo_filename else None
            }
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 400

@app.route('/contact/<int:id>')
def get_contact(id):
    contact = Contact.query.get_or_404(id)
    contact_data = {
        'id': contact.id,
        'name': contact.name,
        'position': contact.position,
        'email': contact.email,
        'phones': json.loads(contact.phones),
        'department': contact.department,
        'notes': contact.notes,
        'photo_url': url_for('static', filename=f'uploads/contacts/{contact.photo_path}') if contact.photo_path else None
    }
    return jsonify(contact_data)

@app.route('/contact/<int:contact_id>', methods=['DELETE'])
def delete_contact(contact_id):
    contact = Contact.query.get_or_404(contact_id)
    db.session.delete(contact)
    db.session.commit()
    return '', 204


@app.errorhandler(404)
def not_found_error(error):
    return render_template('error.html', error=error), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('error.html', error=error), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)